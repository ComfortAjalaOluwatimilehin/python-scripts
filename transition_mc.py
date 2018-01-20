#USAGE: python mc.py <filepaths>

import os, sys, openpyxl, logging
from openpyxl.chart import (
    ScatterChart,
    Reference,
	Series
)
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

def read_each_file(filepath, dictionary ):
		'''
			Reads each file and get series 
			arguments:
				filepath: path of file (string)
				dictionary: dict 
				init: if categories and labels have been setup or not (boolean)
				returns: 
					dictionary (updated)
		'''
		
		wb = openpyxl.load_workbook(filepath)
		ws = wb.active  
		offset = 0
		if type(ws.cell(row = 1, column = 1).value ) == str:
			offset = 1
		
		data = {}
		data["title"] = ws.title
		data["values"] = []
		data["cat"] = []
		if offset > 0:
			#gets x and y  labels 
			data["x"] = ws.cell(row=1, column=1).value 
			data["y"] = ws.title #ws.cell(row=1, column=2).value
		for pos in range(1 + offset, ws.max_row):
			data["cat"].append(  ws.cell(row=pos, column=1).value  )
		for row_pos in range(1 + offset,  ws.max_row ):
			val = ws.cell(row=row_pos,column= 2).value
			data["values"].append(val)
		dictionary["data"].append(data)
		return dictionary
		
def get_data_to_dict(filepaths, target):
	dictionary = {}
	dictionary["data"]  = []
	for filepath in filepaths:
		dictionary = read_each_file(filepath, dictionary)
	wb_path = write_new_workbook(dictionary, target)
	return draw_chart(wb_path)

def write_new_workbook(dictionary: dict, target: str):
	
	wb = openpyxl.Workbook()
	ws = wb.active 
	ws.title = "Multiple Series"

	#write series 
	col = 1
	label_col = 1
	row = 1
	for d in dictionary["data"]:
		#write labels
		ws.cell(row=1, column=label_col, value=d["x"])
		label_col += 1
		ws.cell(row=1, column=label_col, value=d["y"])
		label_col += 1
		for row_pos in range( len(d["cat"]) ):
			cur_row = row_pos + 2
			ws.cell(row=cur_row, column=col, value=d["cat"][row_pos] )
			ws.cell(row=cur_row, column=col + 1, value=d["values"][row_pos] )
			
		col += 2
	path = os.path.join(target, "data_joined.xlsx")
	logging.info("New Excel Sheet created in: " + path)
	wb.save(path)
	return path
		
def draw_chart(wb_path: str):
	wb = openpyxl.load_workbook(wb_path)
	ws = wb.active 
	c1 = ScatterChart()
	c1.y_axis.majorGridlines = None
	c1.x_axis.majorGridlines = None
	c1.x_axis.tickLblSkip = 100
	#data = Reference(ws, min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
	#c1.add_data(data, titles_from_data=True)
	
	
	skips = ws.max_column / 2
	col = 1
	for i in range(1,ws.max_column,2):
		xvalues = Reference(ws, min_col=i, min_row=1, max_row=ws.max_row)
		values = Reference(ws, min_col=i + 1, min_row=1, max_row=ws.max_row)
		series = Series(values, xvalues, title_from_data=True)
		series.smooth = True
		c1.series.append(series)
	

	ws.add_chart(c1, "D10")
	wb.save(wb_path)
	return True
	
	
	
	
	
if __name__ == "__main__":
	if len( sys.argv ) >= 2:
		filepaths = sys.argv[1:]
		valid = True
		#validate paths 
		for filepath in filepaths:
			if  not os.path.exists(filepath):
				valid =False
				break;
		if valid:
			target = filepaths[1]
			target = os.path.dirname(target)
			get_data_to_dict(filepaths, target)
		else:
			logging.error("Some paths are in valid")
			
	else:
		logging.error("USAGE: python mc.py <filepaths>")
		
		
#completed in 01:31:38 YE!!!!