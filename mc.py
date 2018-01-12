#USAGE: python mc.py <filepaths>

import os, sys, openpyxl, logging
from openpyxl.chart import (
    LineChart,
    Reference,
)
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

def read_each_file(filepath, dictionary, init: bool = False ):
		'''
			Reads each file and get series 
			arguments:
				filepath: path of file (string)
				dictionary: dict 
				init: if categories and labels have been setup or not (boolean)
				returns: 
					dictionary (updated)
		'''
		
		wb = openpyxl.load_workbook(filepath)
		ws = wb.active  
		offset = 0
		if type(ws.cell(row = 1, column = 1).value ) == str:
			offset = 1
		if not init:
			if offset > 0:
				#gets x and y  labels 
				dictionary["x"] = ws.cell(row=1, column=1).value 
				dictionary["y"] = ws.cell(row=1, column=2).value
			for pos in range(1 + offset, ws.max_row):
				dictionary["cat"].append(  ws.cell(row=pos, column=1).value  )
		data = {}
		data["title"] = ws.title
		data["values"] = []
		for row_pos in range(1 + offset,  ws.max_row ):
			val = ws.cell(row=row_pos,column= 2).value
			data["values"].append(val)
		dictionary["data"].append(data)
		return dictionary
		
def get_data_to_dict(filepaths, target):
	dictionary = {}
	dictionary["data"]  = []
	dictionary["cat"] = []
	init = False
	for filepath in filepaths:
		dictionary = read_each_file(filepath, dictionary,  init)
		if "cat" in dictionary and len( dictionary["cat"] ) > 0:
			init = True
	wb_path = write_new_workbook(dictionary, target)
	return draw_chart(wb_path)

def write_new_workbook(dictionary: dict, target: str):
	
	wb = openpyxl.Workbook()
	ws = wb.active 
	ws.title = "Multiple Series"
	row = 0
	#write labels 
	if  "x" in dictionary and "y" in dictionary:
		ws.cell(row=1, column=1, value= dictionary["x"])
		ws.cell(row=1, column=2, value=dictionary["y"])
		row += 1
	#write categories 
	i = 0
	for pos in range( len( dictionary["cat"] )  ):
		ws.cell(row= pos + 1 + row , column=1,value= dictionary["cat"][i] )
		i+=1
	#write series 
	colum_pos = 2
	for d in dictionary["data"]:
		ws.cell(row= 1, column= colum_pos, value =d["title"] )
		for pos in range(  len(  d["values"] )  ):
			ws.cell(row= pos + 1 + row, column=colum_pos, value=d["values"][pos] )
		colum_pos+=1
	path = os.path.join(target, "data_joined.xlsx")
	logging.info("New Excel Sheet created in: " + path)
	wb.save(path)
	return path
		
def draw_chart(wb_path: str):
	wb = openpyxl.load_workbook(wb_path)
	ws = wb.active 
	c1 = LineChart()
	c1.y_axis.majorGridlines = None
	c1.x_axis.tickLblSkip = 100
	data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
	c1.add_data(data, titles_from_data=True)
	#assume maximum number of rows = 4; means--> no more than four series in a chart
	symbols = ("triangle", "star", "square", "plus", "dash", "dot")
	#style lines 
	for series_num in range(ws.max_column - 1):
		c1.series[series_num].smooth = True
		#c1.series[series_num].marker.symbol = symbols[series_num]
		#c1.series[series_num].graphicalProperties.line.noFill = True
	ws.add_chart(c1, "D10")
	wb.save(wb_path)
	return True
	
	
	
	
	
if __name__ == "__main__":
	if len( sys.argv ) >= 2:
		filepaths = sys.argv[1:]
		valid = True
		#validate paths 
		for filepath in filepaths:
			if  not os.path.exists(filepath):
				valid =False
				break;
		if valid:
			target = filepaths[1]
			target = os.path.dirname(target)
			get_data_to_dict(filepaths, target)
		else:
			logging.error("Some paths are in valid")
			
	else:
		logging.error("USAGE: python mc.py <filepaths>")
		
		
#completed in 01:31:38 YE!!!!