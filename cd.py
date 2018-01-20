import openpyxl, os

def task(file):
	directory = os.path.dirname(file)
	filename = os.path.basename(file)
	filename = filename.split(".")[0]
	fh = open(os.path.join(directory, file +  ".txt"), "w")
	wb = openpyxl.load_workbook(file)
	ws= wb.active
	counter = 240;
	t  = 0
	for row in range(1, ws.max_row):
		value = ws.cell(row=row+1, column=1).value
		if float(value) == float(counter):
			cd_value = ws.cell(row=row+1, column=2).value
			fh.write(str(cd_value) + "\n")
			counter -= 1
			t+=1
		if t >= 41:
			break;
	fh.close()

	
if __name__ == "__main__":
	dir = "C:\\Users\\ajala\\Downloads\\Research Protocol\\cd_excel"
	for file in os.listdir(dir):
		if file.endswith(".xlsx"):
			task(os.path.join(dir, file))